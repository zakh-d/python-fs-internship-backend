import csv
import datetime
import io
from math import floor
from uuid import UUID

import openpyxl
from pydantic import ValidationError
from sqlalchemy.ext.asyncio import AsyncSession

from app.repositories.company_repository import CompanyRepository
from app.repositories.quizz_repository import QuizzRepository
from app.repositories.user_repository import UserRepository
from app.schemas.quizz_schema import (
    AnswerCreateSchema,
    AnswerSchema,
    AnswerUpdateSchema,
    AnswerWithCorrectSchema,
    ChoosenAnswerDisplaySchema,
    ChoosenAnswerSchema,
    CompletionInfoSchema,
    QuestionCompletionSchema,
    QuestionCreateSchema,
    QuestionResultDisplaySchema,
    QuestionResultSchema,
    QuestionSchema,
    QuestionUpdateSchema,
    QuestionWithCorrectAnswerSchema,
    QuizzCompletionSchema,
    QuizzCreateSchema,
    QuizzDetailResultSchema,
    QuizzListSchema,
    QuizzResultAnalyticsListSchema,
    QuizzResultDisplaySchema,
    QuizzResultDisplayWithUserSchema,
    QuizzResultListDisplaySchema,
    QuizzResultSchema,
    QuizzResultsListForDateSchema,
    QuizzResultWithQuizzDataSchema,
    QuizzResultWithTimestampSchema,
    QuizzResultWithUserSchema,
    QuizzSchema,
    QuizzUpdateSchema,
    QuizzWithCorrectAnswersSchema,
    QuizzWithNoQuestionsSchema,
)
from app.schemas.user_shema import UserDetail, UserSchema
from app.services.notification_service import NotificationService
from app.services.quizz_service.exceptions import QuizzError, QuizzNotFound


class QuizzService:
    def __init__(self, session: AsyncSession) -> None:
        self._quizz_repository = QuizzRepository(session)
        self._company_repository = CompanyRepository(session)
        self._user_repository = UserRepository(session)
        self._notification_service = NotificationService(session)

    async def create_quizz(self, quizz_data: QuizzCreateSchema, company_id: UUID) -> QuizzSchema:
        async with self._quizz_repository.unit():
            quizz = await self._quizz_repository.create_quizz(
                title=quizz_data.title,
                description=quizz_data.description,
                frequency=quizz_data.frequency,
                company_id=company_id,
            )
            response_schema = QuizzSchema(
                **quizz_data.model_dump(exclude={'questions'}),
                questions=[],
                id=quizz.id,
            )
            for question_data in quizz_data.questions:
                question = await self._quizz_repository.create_question(
                    text=question_data.text,
                    quizz_id=quizz.id,
                )
                question_response_schema = QuestionSchema(
                    **question_data.model_dump(exclude={'answers'}),
                    answers=[],
                    id=question.id,
                    multiple=len(list(filter(lambda answer: answer.is_correct, question_data.answers))) > 1,
                )

                for answer_data in question_data.answers:
                    answer = await self._quizz_repository.create_answer(
                        text=answer_data.text,
                        question_id=question.id,
                        is_correct=answer_data.is_correct,
                    )
                    question_response_schema.answers.append(AnswerSchema.model_validate(answer))
                response_schema.questions.append(question_response_schema)

            company = await self._company_repository.get_company_by_id(company_id)
            await self._notification_service.send_notification_to_company_members(
                company_id, 'New quizz', f'New quizz: {quizz.title} has been created in {company.name}'
            )
            return response_schema

    async def add_question_to_quizz(self, quizz_id: UUID, question_data: QuestionCreateSchema) -> None:
        async with self._quizz_repository.unit():
            question = await self._quizz_repository.create_question(
                text=question_data.text,
                quizz_id=quizz_id,
            )
            for answer_data in question_data.answers:
                await self._quizz_repository.create_answer(**answer_data.model_dump(), question_id=question.id)

    async def add_answer_to_question(self, quizz_id: UUID, question_id: UUID, answer_data: AnswerCreateSchema) -> None:
        question = await self._quizz_repository.get_question(question_id)
        if question.quizz_id != quizz_id:
            raise QuizzNotFound('Question')
        if await self._quizz_repository.get_question_answers_count(question_id) == 4:
            raise QuizzError('Question can have max 4 answers')
        async with self._quizz_repository.unit():
            await self._quizz_repository.create_answer(**answer_data.model_dump(), question_id=question_id)

    async def get_quizz(self, quizz_id: UUID) -> QuizzWithNoQuestionsSchema:
        quizz = await self._quizz_repository.get_quizz(quizz_id)
        if not quizz:
            raise QuizzNotFound()
        return QuizzWithNoQuestionsSchema.model_validate(quizz)

    async def fetch_quizz_questions(self, quizz_without_questions: QuizzWithNoQuestionsSchema) -> QuizzSchema:
        quizz = QuizzSchema(**quizz_without_questions.model_dump(), questions=[])
        questions = await self._quizz_repository.get_quizz_questions(quizz.id)
        for question in questions:
            question_schema = QuestionSchema(id=question.id, text=question.text, answers=[], multiple=False)
            answers = await self._quizz_repository.get_question_answers(question.id)
            if len(list(filter(lambda answer: answer.is_correct, answers))) > 1:
                question_schema.multiple = True
            question_schema.answers = [AnswerSchema.model_validate(answer) for answer in answers]
            quizz.questions.append(question_schema)
        return quizz

    async def fetch_quizz_questions_with_correct_answers(
        self, quizz_without_questions: QuizzWithNoQuestionsSchema
    ) -> QuizzWithCorrectAnswersSchema:
        quizz = QuizzSchema(**quizz_without_questions.model_dump(), questions=[])
        questions = await self._quizz_repository.get_quizz_questions(quizz.id)
        for question in questions:
            question_schema = QuestionWithCorrectAnswerSchema(id=question.id, text=question.text, answers=[])
            answers = await self._quizz_repository.get_question_answers(question.id)
            question_schema.answers = [AnswerWithCorrectSchema.model_validate(answer) for answer in answers]
            quizz.questions.append(question_schema)
        return quizz

    async def get_company_quizzes(self, company_id: UUID, page: int, limit: int) -> QuizzListSchema:
        offset = (page - 1) * limit
        quizzes = await self._quizz_repository.get_company_quizzes(company_id, offset, limit)
        return QuizzListSchema(
            quizzes=[QuizzWithNoQuestionsSchema.model_validate(quizz) for quizz in quizzes],
            total_count=await self._quizz_repository.get_company_quizzes_count(company_id),
        )

    async def delete_quizz(self, quizz_id: UUID) -> None:
        await self._quizz_repository.delete_quizz_and_commit(quizz_id)

    async def delete_question(self, question_id: UUID, quizz_id: UUID) -> None:
        if await self._quizz_repository.get_quizz_questions_count(quizz_id) < 2:
            raise QuizzError('Cannot delete last question')
        question = await self._quizz_repository.delete_question_and_commit(question_id)
        if not question:
            raise QuizzNotFound('Question')
        if question.quizz_id != quizz_id:
            raise QuizzNotFound()
        await self._quizz_repository.delete_question_and_commit(question_id)

    async def delete_answer(self, answer_id: UUID, quizz_id: UUID) -> None:
        answer = await self._quizz_repository.get_answer(answer_id)
        if not answer:
            raise QuizzNotFound('Answer')
        question = await self._quizz_repository.get_question(answer.question_id)
        if await self._quizz_repository.get_question_answers_count(question.id) < 3:
            raise QuizzError('Question must have at least two answer')
        if answer.is_correct:
            answers = await self._quizz_repository.get_question_answers(question.id)
            correct_answers = [answer for answer in answers if answer.is_correct]
            if len(correct_answers) < 2:
                raise QuizzError('Cannot delete only correct answer')
        if question.quizz_id != quizz_id:
            raise QuizzNotFound()
        await self._quizz_repository.delete_answer_and_commit(answer_id)

    async def update_quizz(self, quizz_id: UUID, quizz_data: QuizzUpdateSchema) -> QuizzWithNoQuestionsSchema:
        quizz = await self._quizz_repository.get_quizz(quizz_id)
        if not quizz:
            raise QuizzNotFound()
        async with self._quizz_repository.unit():
            quizz = await self._quizz_repository.update_quizz(quizz, quizz_data)
        return QuizzWithNoQuestionsSchema.model_validate(quizz)

    async def update_question(self, question_id: UUID, quizz_id: UUID, question_data: QuestionUpdateSchema) -> None:
        question = await self._quizz_repository.get_question(question_id)
        if not question:
            raise QuizzNotFound('Question')
        if question.quizz_id != quizz_id:
            raise QuizzNotFound('Question')
        async with self._quizz_repository.unit():
            await self._quizz_repository.update_question(question, question_data)

    async def update_answer(self, answer_id: UUID, quizz_id: UUID, answer_data: AnswerUpdateSchema) -> None:
        answer = await self._quizz_repository.get_answer(answer_id)
        if not answer:
            raise QuizzNotFound('Answer')
        question = await self._quizz_repository.get_question(answer.question_id)
        if question.quizz_id != quizz_id:
            raise QuizzNotFound('Answer')
        if answer.is_correct and not answer_data.is_correct:
            answers = await self._quizz_repository.get_question_answers(question.id)
            correct_answers = [answer for answer in answers if answer.is_correct]
            if len(correct_answers) < 2:
                raise QuizzError('Question must have at least one correct answers')
        async with self._quizz_repository.unit():
            await self._quizz_repository.update_answer(answer, answer_data)

    async def evaluate_question(
        self, quizz_id: UUID, question_data: QuestionCompletionSchema
    ) -> tuple[float, QuestionResultSchema]:
        question = await self._quizz_repository.get_question(question_data.question_id)
        if not question:
            raise QuizzNotFound('Question')
        if question.quizz_id != quizz_id:
            raise QuizzNotFound('Question')
        answers = await self._quizz_repository.get_question_answers(question.id)
        correct_answers_count = len([answer for answer in answers if answer.is_correct])
        correct_responses = 0
        result = QuestionResultSchema(question_id=question.id, choosen_answers=[])
        for answer_id in question_data.answer_ids:
            answer = await self._quizz_repository.get_answer(answer_id)
            if not answer:
                raise QuizzNotFound('Answer')
            if answer.question_id != question.id:
                raise QuizzNotFound('Answer')
            if answer.is_correct:
                result.choosen_answers.append(ChoosenAnswerSchema(is_correct=True, answer_id=answer.id))
                correct_responses += 1
            else:
                result.choosen_answers.append(ChoosenAnswerSchema(is_correct=False, answer_id=answer.id))
                correct_responses -= 1
        if any(not answer.is_correct for answer in result.choosen_answers):
            correct_responses = 0
        else:
            correct_responses = max(0, correct_responses)
        return correct_responses / correct_answers_count, result

    async def evaluate_quizz(
        self, quizz: QuizzWithNoQuestionsSchema, data: QuizzCompletionSchema, user: UserDetail
    ) -> QuizzResultSchema:
        question_count = await self._quizz_repository.get_quizz_questions_count(data.quizz_id)
        score = 0
        asssesment = QuizzDetailResultSchema(quizz_id=quizz.id, user_id=user.id, questions=[])
        for question in data.questions:
            question_score, question_result = await self.evaluate_question(data.quizz_id, question)
            score += question_score / question_count
            asssesment.questions.append(question_result)
        result = await self._quizz_repository.create_quizz_result(
            user_id=user.id,
            quizz_id=data.quizz_id,
            company_id=quizz.company_id,
            score=floor(score * 100),
        )
        await self._quizz_repository.commit()
        await self._quizz_repository.delete_cached_quizz_for_user(user.id, data.quizz_id)
        await self._quizz_repository.cache_quizz_result(
            user_id=user.id, company_id=quizz.company_id, quizz_id=data.quizz_id, data=asssesment
        )
        return QuizzResultSchema(score=result.score)

    async def get_average_score_by_company(self, company_id: UUID) -> QuizzResultSchema:
        return QuizzResultSchema(score=await self._quizz_repository.get_average_score_by_company(company_id))

    async def get_average_score_by_user(self, user_id: UUID) -> QuizzResultSchema:
        return QuizzResultSchema(score=await self._quizz_repository.get_average_score_by_user(user_id))

    async def get_average_score_for_user_by_quizzes_over_intervals(
        self, user_id: UUID, interval: datetime.timedelta
    ) -> list[QuizzResultAnalyticsListSchema]:
        end_date = datetime.datetime.now()  # noqa: DTZ005
        total_results = []
        while True:
            data = await self._quizz_repository.get_average_score_by_user_grouped_by_quizz_within_dates(
                user_id, datetime.datetime.min, end_date
            )
            if len(data) == 0:
                break
            total_results.append(
                QuizzResultAnalyticsListSchema(
                    results=[
                        QuizzResultWithQuizzDataSchema(
                            quizz_id=result.quizz_id,
                            score=result.average_score,
                            quizz_title=result.title,
                        )
                        for result in data
                    ],
                    date=end_date,
                )
            )
            end_date -= interval
        return total_results

    async def get_lastest_user_completions(self, user_id: UUID) -> list[CompletionInfoSchema]:
        completions = await self._quizz_repository.get_lastest_user_completion_across_all_quizzes(user_id)
        return [
            CompletionInfoSchema(
                quizz_id=completion.quizz_id,
                quizz_title=completion.title,
                completion_time=completion.lastest_completion,
            )
            for completion in completions
        ]

    async def get_user_quizz_completions(self, user_id: UUID, quizz_id: UUID) -> list[QuizzResultWithTimestampSchema]:
        complitions = await self._quizz_repository.get_user_quizz_completions(user_id, quizz_id)
        return [
            QuizzResultWithTimestampSchema(
                score=complition.score,
                completion_time=complition.created_at,
            )
            for complition in complitions
        ]

    async def get_average_scores_for_company_members_over_intervals(
        self, company_id: UUID, interval: datetime.timedelta
    ) -> list[QuizzResultsListForDateSchema]:
        end_date = datetime.datetime.now()  # noqa: DTZ005
        total_results = []
        while True:
            data = await self._quizz_repository.get_average_score_for_company_members_within_dates(
                company_id, datetime.datetime.min, end_date
            )
            if len(data) == 0:
                break
            total_results.append(
                QuizzResultsListForDateSchema(
                    results=[
                        QuizzResultWithUserSchema(
                            score=result.average_score,
                            user_email=result.email,
                        )
                        for result in data
                    ],
                    date=end_date,
                )
            )
            end_date -= interval
        return total_results

    async def get_average_scores_for_quizz_completed_by_user_over_intervals(
        self, user_id: UUID, quizz_id: UUID, interval: datetime.timedelta
    ) -> list[QuizzResultWithTimestampSchema]:
        end_date = datetime.datetime.now()  # noqa: DTZ005
        total_results = []
        while True:
            avg = await self._quizz_repository.get_average_score_for_quizz_by_user(
                quizz_id=quizz_id, user_id=user_id, start_date=datetime.datetime.min, end_date=end_date
            )
            if avg is None:
                break
            total_results.append(QuizzResultWithTimestampSchema(score=avg, completion_time=end_date))
            end_date -= interval
        return total_results

    async def get_average_score_by_quizz(self, quizz_id: UUID) -> QuizzResultSchema:
        return QuizzResultSchema(score=await self._quizz_repository.get_average_score_by_quizz(quizz_id))

    async def get_quizz_response_displayed(self, response: QuizzDetailResultSchema) -> QuizzResultDisplayWithUserSchema:
        quizz = await self.get_quizz(response.quizz_id)
        quizz = await self.fetch_quizz_questions(quizz)
        quizz_result = await self._quizz_repository.get_latest_quizz_result(response.user_id, response.quizz_id)
        if not quizz_result:
            raise QuizzNotFound('Quizz result')

        questions_text_by_ids = {question.id: question.text for question in quizz.questions}
        answers_text_by_ids = {answer.id: answer.text for question in quizz.questions for answer in question.answers}
        user = await self._user_repository.get_user_by_id(response.user_id)
        result_display = QuizzResultDisplayWithUserSchema(score=quizz_result.score, user_email=user.email, questions=[])

        for question in response.questions:
            question_text = questions_text_by_ids[question.question_id]
            choosen_answers = []
            for choosen_answer in question.choosen_answers:
                answer_text = answers_text_by_ids[choosen_answer.answer_id]
                choosen_answers.append(
                    ChoosenAnswerDisplaySchema(text=answer_text, is_correct=choosen_answer.is_correct)
                )
            result_display.questions.append(
                QuestionResultDisplaySchema(text=question_text, choosen_answers=choosen_answers)
            )

        return result_display

    async def get_cached_user_response_json(self, user_id: UUID, quizz_id: UUID) -> QuizzResultDisplaySchema:
        result = await self._quizz_repository.get_user_quizz_response_from_cache(user_id, quizz_id)
        if result is None:
            raise QuizzNotFound('User response')
        return await self.get_quizz_response_displayed(result)

    async def get_cached_user_response_csv(self, user_id: UUID, quizz_id: UUID) -> str:
        response = await self._quizz_repository.get_user_quizz_response_from_cache(user_id, quizz_id)
        if response is None:
            raise QuizzNotFound('User response')

        response_displayed = await self.get_quizz_response_displayed(response)

        output = io.StringIO()
        writter = csv.DictWriter(output, fieldnames=['Question', 'Answer', 'Is Correct'])
        writter.writeheader()
        for question in response_displayed.questions:
            for choosen_answer in question.choosen_answers:
                writter.writerow(
                    {
                        'Question': question.text,
                        'Answer': choosen_answer.text,
                        'Is Correct': choosen_answer.is_correct,
                    }
                )
        return output.getvalue()

    async def get_cached_users_responses_json(
        self, users: list[UserSchema], quizz_id: UUID
    ) -> QuizzResultListDisplaySchema:
        list_ = QuizzResultListDisplaySchema(responses=[])
        for user in users:
            try:
                response = await self.get_cached_user_response_json(user.id, quizz_id)
                list_.responses.append(
                    QuizzResultDisplayWithUserSchema(
                        score=response.score, questions=response.questions, user_email=user.email
                    )
                )
            except QuizzNotFound:
                continue
        return list_

    async def get_cached_users_responses_csv(self, users: list[UserSchema], quizz_id: UUID) -> str:
        output = io.StringIO()
        writter = csv.DictWriter(output, fieldnames=['User', 'Question', 'Answer', 'Is Correct'])
        writter.writeheader()
        for user in users:
            try:
                response = await self.get_cached_user_response_json(user.id, quizz_id)
                for question in response.questions:
                    for choosen_answer in question.choosen_answers:
                        writter.writerow(
                            {
                                'User': user.email,
                                'Question': question.text,
                                'Answer': choosen_answer.text,
                                'Is Correct': choosen_answer.is_correct,
                            }
                        )
            except QuizzNotFound:
                continue
        return output.getvalue()

    async def _user_responses_to_displayed_json(
        self, responses: list[QuizzDetailResultSchema]
    ) -> QuizzResultListDisplaySchema:
        list_ = QuizzResultListDisplaySchema(responses=[])
        for response in responses:
            list_.responses.append(await self.get_quizz_response_displayed(response))
        return list_

    async def get_user_responses_from_cache_json(self, user_id: UUID) -> QuizzResultListDisplaySchema:
        responses = await self._quizz_repository.get_user_cached_responses(user_id)
        return await self._user_responses_to_displayed_json(responses)

    async def get_user_responses_in_company_from_cache_json(
        self, user_id: UUID, company_id: UUID
    ) -> QuizzResultListDisplaySchema:
        responses = await self._quizz_repository.get_user_cached_responses_in_company(user_id, company_id)
        return await self._user_responses_to_displayed_json(responses)

    async def _user_responses_to_displayed_csv(self, responses: list[QuizzDetailResultSchema]) -> str:
        output = io.StringIO()
        writter = csv.DictWriter(output, fieldnames=['Quizz', 'User', 'Question', 'Answer', 'Is Correct'])
        writter.writeheader()
        for response in responses:
            response_displayed = await self.get_quizz_response_displayed(response)
            for question in response_displayed.questions:
                for choosen_answer in question.choosen_answers:
                    writter.writerow(
                        {
                            'Quizz': (await self._quizz_repository.get_quizz(response.quizz_id)).title,
                            'User': response_displayed.user_email,
                            'Question': question.text,
                            'Answer': choosen_answer.text,
                            'Is Correct': choosen_answer.is_correct,
                        }
                    )
        return output.getvalue()

    async def get_user_responses_from_cache_csv(self, user_id: UUID) -> str:
        responses = await self._quizz_repository.get_user_cached_responses(user_id)
        return await self._user_responses_to_displayed_csv(responses)

    async def get_user_responses_in_company_from_cache_csv(self, user_id: UUID, company_id: UUID) -> str:
        responses = await self._quizz_repository.get_user_cached_responses_in_company(user_id, company_id)
        return await self._user_responses_to_displayed_csv(responses)

    async def get_company_members_responses_from_cache_json(self, company_id: UUID) -> QuizzResultListDisplaySchema:
        responses = await self._quizz_repository.get_company_members_responses(company_id)
        return await self._user_responses_to_displayed_json(responses)

    async def get_company_members_responses_from_cache_csv(self, company_id: UUID) -> str:
        responses = await self._quizz_repository.get_company_members_responses(company_id)
        return await self._user_responses_to_displayed_csv(responses)

    def get_schema_from_excel(self, file: bytes, company_id: UUID) -> QuizzCreateSchema:
        workbook = openpyxl.load_workbook(io.BytesIO(file))
        ws = workbook.active

        excel_quizz_prolog = ['QUIZZ TITLE:', 'QUIZZ DESCRIPTION:', 'QUIZZ FREQUENCY:', 'QUESTION:']

        for i, expected_value in enumerate(excel_quizz_prolog):
            if ws[f'A{i + 1}'].value != expected_value:
                raise QuizzError(
                    'Invalid file format, download example to see correct format\n'
                    f'Expected: {expected_value}, got: {ws[f"A{i + 1}"].value}'
                )

        quizz_title = str(ws['B1'].value)
        quizz_description = str(ws['B2'].value)
        frequency = int(ws['B3'].value)

        questions = []
        curr_row = 4
        try:
            while ws[f'A{curr_row}'].value == 'QUESTION:':
                question_text = str(ws[f'B{curr_row}'].value)
                curr_row += 1

                answers = []

                while ws[f'A{curr_row}'].value == 'ANSWER:':
                    answer_text = str(ws[f'B{curr_row}'].value)
                    is_correct = ws[f'C{curr_row}'].value == 'CORRECT'
                    answers.append(AnswerCreateSchema(text=answer_text, is_correct=is_correct))
                    curr_row += 1

                question_data = QuestionCreateSchema(text=question_text, answers=answers)
                questions.append(question_data)

            quizz_data = QuizzCreateSchema(
                company_id=company_id,
                title=quizz_title,
                description=quizz_description,
                frequency=frequency,
                questions=questions,
            )
            return quizz_data
        except ValidationError as e:
            raise QuizzError(e.errors()[0]['msg'])

    async def create_or_update_quizz(self, quizz_schema: QuizzCreateSchema) -> QuizzSchema:
        quizz = await self._quizz_repository.get_quizz_by_company_and_title(quizz_schema.company_id, quizz_schema.title)

        if not quizz:
            return await self.create_quizz(quizz_schema, quizz_schema.company_id)

        async with self._quizz_repository.unit():
            await self._quizz_repository.update_quizz(
                quizz,
                QuizzUpdateSchema(
                    title=quizz_schema.title, description=quizz_schema.description, frequency=quizz_schema.frequency
                ),
            )
            quizz_questions = await self._quizz_repository.get_quizz_questions(quizz.id)
            new_questions_text = [question.text for question in quizz_schema.questions]

            for question in quizz_questions:
                if question.text not in new_questions_text:
                    await self.delete_question(question.id)

            for question_schema in quizz_schema.questions:
                question = await self._quizz_repository.get_question_by_quizz_and_text(quizz.id, question_schema.text)
                if not question:
                    await self.add_question_to_quizz(quizz.id, question_schema)
                    continue

                await self.update_question(question.id, quizz.id, QuestionUpdateSchema(text=question_schema.text))

                new_answers_text = [answer.text for answer in question_schema.answers]
                current_answers = await self._quizz_repository.get_question_answers(question.id)
                current_answers_text = [answer.text for answer in current_answers]

                for answer in current_answers:
                    if answer.text not in new_answers_text:
                        current_answers_text.remove(answer.text)
                        await self._quizz_repository.delete_answer(answer.id)

                for answer_schema in question_schema.answers:
                    if answer_schema.text not in current_answers_text:
                        await self.add_answer_to_question(quizz.id, question.id, answer_schema)
                        continue
                    answer = await self._quizz_repository.get_answer_by_question_and_text(
                        question.id, answer_schema.text
                    )
                    await self.update_answer(
                        answer.id,
                        quizz.id,
                        AnswerUpdateSchema(text=answer_schema.text, is_correct=answer_schema.is_correct),
                    )
        return await self.fetch_quizz_questions(await self.get_quizz(quizz.id))
